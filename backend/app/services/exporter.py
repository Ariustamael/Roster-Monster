"""
Excel export service — generates roster spreadsheets in two formats:
  1. Full: complete multi-sheet workbook (daily roster, clinic, unavailability)
  2. Clean: optimized sheets for distribution (call sheet, clinic, combined views)
"""

import calendar
from collections import defaultdict
from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from ..models import (
    CallAssignment,
    DutyAssignment,
    Staff,
    ConsultantOnCall,
    ACOnCall,
    DutyType,
    CallTypeConfig,
    PublicHoliday,
    Leave,
    MonthlyConfig,
)


# ── Shared styles ────────────────────────────────────────────────────────

HEADER_FONT = Font(bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14)
HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
WEEKEND_FILL = PatternFill("solid", fgColor="FFF2CC")
PH_FILL = PatternFill("solid", fgColor="FCE4EC")
CLINIC_FILL = PatternFill("solid", fgColor="D1FAE5")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
CENTER = Alignment(horizontal="center", vertical="center")
WRAP = Alignment(wrap_text=True, vertical="top")

DATA_FONT = Font(size=9)
OVERRIDE_FILL = PatternFill(
    "solid", fgColor="FFE0B2"
)  # light orange for manual overrides


def _set_col_width(ws, col: int, width: int) -> int:
    """Set column width and return the next column index."""
    ws.column_dimensions[get_column_letter(col)].width = width
    return col + 1


def _style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def _style_data_cell(ws, row, col, is_weekend=False, is_ph=False):
    cell = ws.cell(row=row, column=col)
    cell.border = THIN_BORDER
    if is_ph:
        cell.fill = PH_FILL
    elif is_weekend:
        cell.fill = WEEKEND_FILL


# ── Shared helpers ───────────────────────────────────────────────────────


def _apply_print_settings(ws):
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)
    ws.print_title_rows = "1:3"


async def _get_call_type_columns(db) -> list[str]:
    result = await db.execute(
        select(CallTypeConfig)
        .filter(CallTypeConfig.is_active.is_(True))
        .order_by(CallTypeConfig.display_order)
    )
    return [ct.name for ct in result.scalars().all()]


async def _get_ph_dates(db, year: int, month: int) -> set[date]:
    import calendar as _cal

    last_day = _cal.monthrange(year, month)[1]
    result = await db.execute(
        select(PublicHoliday).filter(
            PublicHoliday.date >= date(year, month, 1),
            PublicHoliday.date <= date(year, month, last_day),
        )
    )
    return {r.date for r in result.scalars().all()}


# ── Entry points ─────────────────────────────────────────────────────────


async def export_full(config: MonthlyConfig, db: AsyncSession) -> BytesIO:
    wb = Workbook()
    year, month = config.year, config.month
    month_name = calendar.month_name[month]
    num_days = calendar.monthrange(year, month)[1]

    await _build_daily_roster_full(wb, config, db, year, month, num_days, month_name)
    await _build_clinic_full(wb, config, db, year, month, num_days, month_name)
    await _build_unavailability(wb, config, db, year, month, month_name)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


async def export_clean(config: MonthlyConfig, db: AsyncSession) -> BytesIO:
    wb = Workbook()
    year, month = config.year, config.month
    month_name = calendar.month_name[month]
    num_days = calendar.monthrange(year, month)[1]

    await _build_call_sheet_clean(wb, config, db, year, month, num_days, month_name)
    await _build_clinic_sheet_clean(wb, config, db, year, month, num_days, month_name)
    await _build_combined_call_ot(wb, config, db, year, month, num_days, month_name)
    await _build_combined_ot_clinic(wb, config, db, year, month, num_days, month_name)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Builder stubs (bodies implemented in later tasks) ────────────────────


async def _build_daily_roster_full(
    wb: Workbook,
    config: MonthlyConfig,
    db: AsyncSession,
    year: int,
    month: int,
    num_days: int,
    month_name: str,
):
    ws = wb.active
    ws.title = "Daily Roster"
    year_2digit = str(year)[-2:]

    # ── Fetch all data ────────────────────────────────────────────────────

    call_type_cols = await _get_call_type_columns(db)
    ph_dates = await _get_ph_dates(db, year, month)

    result = await db.execute(
        select(CallAssignment)
        .filter(CallAssignment.config_id == config.id)
        .options(selectinload(CallAssignment.staff))
    )
    calls = result.scalars().all()

    result = await db.execute(
        select(ConsultantOnCall)
        .filter(ConsultantOnCall.config_id == config.id)
        .options(
            selectinload(ConsultantOnCall.consultant),
            selectinload(ConsultantOnCall.supervising_consultant),
        )
    )
    consultants_oc = result.scalars().all()

    result = await db.execute(
        select(ACOnCall)
        .filter(ACOnCall.config_id == config.id)
        .options(selectinload(ACOnCall.ac))
    )
    acs_oc = result.scalars().all()

    result = await db.execute(
        select(DutyAssignment)
        .filter(
            DutyAssignment.config_id == config.id,
            DutyAssignment.duty_type.in_(
                [DutyType.OT, DutyType.WARD_MO, DutyType.EOT_MO, DutyType.EOT]
            ),
        )
        .options(selectinload(DutyAssignment.staff))
    )
    duties = result.scalars().all()

    # ── Index data by date ────────────────────────────────────────────────

    # consultant on-call: date -> ConsultantOnCall
    consultant_by_date: dict[date, ConsultantOnCall] = {}
    for c in consultants_oc:
        consultant_by_date[c.date] = c

    # AC on-call: date -> ACOnCall
    ac_by_date: dict[date, ACOnCall] = {}
    for a in acs_oc:
        ac_by_date[a.date] = a

    # call assignments: (date, call_type) -> list[CallAssignment]

    call_by_date_type: dict[tuple, list] = defaultdict(list)
    for ca in calls:
        call_by_date_type[(ca.date, ca.call_type)].append(ca)

    # duty assignments by date and type
    ot_by_date: dict[date, dict[str, list[str]]] = defaultdict(
        lambda: defaultdict(list)
    )  # date -> location -> [names]
    ward_mo_by_date: dict[date, list[str]] = defaultdict(list)
    eot_mo_by_date: dict[date, list[str]] = defaultdict(list)
    eot_by_date: dict[date, list[str]] = defaultdict(list)

    # Collect unique OT locations in encounter order
    ot_locations: list[str] = []
    ot_locations_seen: set[str] = set()

    for d in duties:
        staff_name = d.staff.name if d.staff else ""
        if d.duty_type == DutyType.OT:
            loc = d.location or "?"
            if loc not in ot_locations_seen:
                ot_locations.append(loc)
                ot_locations_seen.add(loc)
            ot_by_date[d.date][loc].append(staff_name)
        elif d.duty_type == DutyType.WARD_MO:
            ward_mo_by_date[d.date].append(staff_name)
        elif d.duty_type == DutyType.EOT_MO:
            eot_mo_by_date[d.date].append(staff_name)
        elif d.duty_type == DutyType.EOT:
            eot_by_date[d.date].append(staff_name)

    # ── Build column layout ───────────────────────────────────────────────

    fixed_headers = (
        ["Date", "Day", "Consultant", "AC"]
        + call_type_cols
        + ["Ward MO", "EOT MO", "EOT"]
        + ot_locations
    )
    num_cols = len(fixed_headers)

    # ── Title row ─────────────────────────────────────────────────────────

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    title_cell = ws.cell(
        row=1, column=1, value=f"Daily Roster - {month_name}'{year_2digit}"
    )
    title_cell.font = TITLE_FONT
    title_cell.alignment = CENTER

    # Row 2: spacer (leave empty)

    # ── Header row (row 3) ────────────────────────────────────────────────

    for col_idx, header in enumerate(fixed_headers, start=1):
        ws.cell(row=3, column=col_idx, value=header)
    _style_header_row(ws, 3, num_cols)

    # ── Data rows ─────────────────────────────────────────────────────────

    day_abbrevs = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]

    for day in range(1, num_days + 1):
        row = day + 3  # rows 4+
        current_date = date(year, month, day)
        weekday = current_date.weekday()  # 0=Mon, 5=Sat, 6=Sun
        is_weekend = weekday >= 5
        is_ph = current_date in ph_dates

        day_label = day_abbrevs[weekday]
        if is_ph:
            day_label += " (PH)"

        # Consultant label
        cons_label = ""
        if current_date in consultant_by_date:
            coc = consultant_by_date[current_date]
            cons_name = coc.consultant.name if coc.consultant else ""
            if coc.supervising_consultant_id and coc.supervising_consultant:
                cons_label = f"{cons_name} / {coc.supervising_consultant.name}"
            else:
                cons_label = cons_name

        # AC label
        ac_label = ""
        if current_date in ac_by_date:
            aoc = ac_by_date[current_date]
            ac_label = aoc.ac.name if aoc.ac else ""

        row_values = [day, day_label, cons_label, ac_label]

        # Call type columns
        call_overrides: list[bool] = []
        for ct_name in call_type_cols:
            assignments = call_by_date_type.get((current_date, ct_name), [])
            if assignments:
                names = ", ".join(a.staff.name for a in assignments if a.staff)
                is_override = any(a.is_manual_override for a in assignments)
            else:
                names = ""
                is_override = False
            row_values.append(names)
            call_overrides.append(is_override)

        # Ward MO, EOT MO, EOT
        row_values.append(", ".join(ward_mo_by_date.get(current_date, [])))
        row_values.append(", ".join(eot_mo_by_date.get(current_date, [])))
        row_values.append(", ".join(eot_by_date.get(current_date, [])))

        # OT room columns
        for loc in ot_locations:
            names = ", ".join(ot_by_date[current_date].get(loc, []))
            row_values.append(names)

        # Write cells
        for col_idx, value in enumerate(row_values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = CENTER
            if is_ph:
                cell.fill = PH_FILL
            elif is_weekend:
                cell.fill = WEEKEND_FILL

        # Apply override fill to call type cells (overrides weekend/PH)
        for i, is_override in enumerate(call_overrides):
            if is_override:
                col_idx = 5 + i  # call type cols start at column 5 (1-indexed)
                ws.cell(row=row, column=col_idx).fill = OVERRIDE_FILL

    # ── Column widths ─────────────────────────────────────────────────────

    col = 1
    col = _set_col_width(ws, col, 5)  # Date
    col = _set_col_width(ws, col, 8)  # Day
    col = _set_col_width(ws, col, 18)  # Consultant
    col = _set_col_width(ws, col, 14)  # AC
    for _ in call_type_cols:
        col = _set_col_width(ws, col, 12)
    col = _set_col_width(ws, col, 14)  # Ward MO
    col = _set_col_width(ws, col, 14)  # EOT MO
    col = _set_col_width(ws, col, 14)  # EOT
    for _ in ot_locations:
        col = _set_col_width(ws, col, 14)


async def _build_clinic_full(
    wb: Workbook,
    config: MonthlyConfig,
    db: AsyncSession,
    year: int,
    month: int,
    num_days: int,
    month_name: str,
):
    ws = wb.create_sheet("Clinic")
    year_2digit = str(year)[-2:]

    ph_dates = await _get_ph_dates(db, year, month)

    result = await db.execute(
        select(DutyAssignment)
        .filter(
            DutyAssignment.config_id == config.id,
            DutyAssignment.duty_type == DutyType.CLINIC,
        )
        .options(selectinload(DutyAssignment.staff))
        .order_by(DutyAssignment.date, DutyAssignment.session, DutyAssignment.location)
    )
    clinic_duties = result.scalars().all()

    # ── Discover clinic column keys in encounter order ────────────────────

    clinic_cols: list[str] = []
    clinic_cols_seen: set[str] = set()

    for d in clinic_duties:
        key = d.clinic_type if d.clinic_type else (d.location if d.location else "?")
        if key not in clinic_cols_seen:
            clinic_cols.append(key)
            clinic_cols_seen.add(key)

    # ── Group duties by (date, session) ──────────────────────────────────

    # (date, session) -> clinic_key -> [staff_name]
    clinic_by_date_session: dict[tuple, dict[str, list[str]]] = defaultdict(
        lambda: defaultdict(list)
    )
    date_sessions_ordered: list[tuple] = []
    date_sessions_seen: set[tuple] = set()

    for d in clinic_duties:
        key = d.clinic_type if d.clinic_type else (d.location if d.location else "?")
        staff_name = d.staff.name if d.staff else ""
        ds = (d.date, d.session)
        if ds not in date_sessions_seen:
            date_sessions_ordered.append(ds)
            date_sessions_seen.add(ds)
        clinic_by_date_session[ds][key].append(staff_name)

    # ── Build column layout ───────────────────────────────────────────────

    fixed_headers = ["Date", "Day", "Session"] + clinic_cols
    num_cols = len(fixed_headers)

    # ── Title row ─────────────────────────────────────────────────────────

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(num_cols, 1))
    title_cell = ws.cell(row=1, column=1, value=f"Clinic - {month_name}'{year_2digit}")
    title_cell.font = TITLE_FONT
    title_cell.alignment = CENTER

    # Row 2: spacer

    # ── Header row (row 3) ────────────────────────────────────────────────

    for col_idx, header in enumerate(fixed_headers, start=1):
        ws.cell(row=3, column=col_idx, value=header)
    _style_header_row(ws, 3, num_cols)

    # ── Data rows ─────────────────────────────────────────────────────────

    day_abbrevs = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]

    for data_row_idx, (current_date, session) in enumerate(date_sessions_ordered):
        row = data_row_idx + 4
        weekday = current_date.weekday()
        is_ph = current_date in ph_dates

        day_label = day_abbrevs[weekday]
        # Session enum value as string
        session_label = session.value if hasattr(session, "value") else str(session)

        row_values = [current_date.day, day_label, session_label]
        for ck in clinic_cols:
            names = clinic_by_date_session[(current_date, session)].get(ck, [])
            row_values.append(", ".join(names) if names else "")

        for col_idx, value in enumerate(row_values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = CENTER
            # PH overrides clinic_fill; clinic_fill overrides weekend
            if is_ph:
                cell.fill = PH_FILL
            else:
                cell.fill = CLINIC_FILL

    # ── Column widths ─────────────────────────────────────────────────────

    col = 1
    col = _set_col_width(ws, col, 5)  # Date
    col = _set_col_width(ws, col, 7)  # Day
    col = _set_col_width(ws, col, 8)  # Session
    for _ in clinic_cols:
        col = _set_col_width(ws, col, 12)


async def _build_unavailability(
    wb: Workbook,
    config: MonthlyConfig,
    db: AsyncSession,
    year: int,
    month: int,
    month_name: str,
):
    ws = wb.create_sheet("Unavailability")
    year_2digit = str(year)[-2:]
    num_days = calendar.monthrange(year, month)[1]

    result = await db.execute(
        select(Leave)
        .join(Leave.staff)
        .filter(
            Leave.date >= date(year, month, 1),
            Leave.date <= date(year, month, num_days),
        )
        .options(selectinload(Leave.staff))
        .order_by(Leave.date, Staff.name)
    )
    leaves = result.scalars().all()

    # ── Build column layout ───────────────────────────────────────────────

    headers = ["Date", "Day", "Staff Name", "Rank", "Leave Type"]
    num_cols = len(headers)

    # ── Title row ─────────────────────────────────────────────────────────

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    title_cell = ws.cell(
        row=1, column=1, value=f"Unavailability - {month_name}'{year_2digit}"
    )
    title_cell.font = TITLE_FONT
    title_cell.alignment = CENTER

    # Row 2: spacer

    # ── Header row (row 3) ────────────────────────────────────────────────

    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=3, column=col_idx, value=header)
    _style_header_row(ws, 3, num_cols)

    # ── Data rows ─────────────────────────────────────────────────────────

    day_abbrevs = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]

    for data_row_idx, leave in enumerate(leaves):
        row = data_row_idx + 4
        current_date = leave.date
        weekday = current_date.weekday()
        day_label = day_abbrevs[weekday]

        staff_name = leave.staff.name if leave.staff else ""
        rank = leave.staff.rank if leave.staff else ""
        leave_type = leave.leave_type if leave.leave_type else ""

        row_data = [
            (current_date.day, True),  # Date — centered
            (day_label, True),  # Day — centered
            (staff_name, False),  # Staff Name — left-aligned
            (rank, False),  # Rank — left-aligned
            (leave_type, False),  # Leave Type — left-aligned
        ]

        for col_idx, (value, center_align) in enumerate(row_data, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = (
                CENTER
                if center_align
                else Alignment(horizontal="left", vertical="center")
            )

    # ── Column widths ─────────────────────────────────────────────────────

    ws.column_dimensions[get_column_letter(1)].width = 8  # Date
    ws.column_dimensions[get_column_letter(2)].width = 7  # Day
    ws.column_dimensions[get_column_letter(3)].width = 20  # Staff Name
    ws.column_dimensions[get_column_letter(4)].width = 18  # Rank
    ws.column_dimensions[get_column_letter(5)].width = 14  # Leave Type


async def _build_call_sheet_clean(
    wb: Workbook,
    config: MonthlyConfig,
    db: AsyncSession,
    year: int,
    month: int,
    num_days: int,
    month_name: str,
):
    ws = wb.active
    ws.title = "Call Roster"
    year_2digit = str(year)[-2:]
    _apply_print_settings(ws)

    call_type_cols = await _get_call_type_columns(db)
    ph_dates = await _get_ph_dates(db, year, month)

    result = await db.execute(
        select(CallAssignment)
        .filter(CallAssignment.config_id == config.id)
        .options(selectinload(CallAssignment.staff))
    )
    calls = result.scalars().all()

    result = await db.execute(
        select(ConsultantOnCall)
        .filter(ConsultantOnCall.config_id == config.id)
        .options(
            selectinload(ConsultantOnCall.consultant),
            selectinload(ConsultantOnCall.supervising_consultant),
        )
    )
    consultants_oc = result.scalars().all()

    result = await db.execute(
        select(ACOnCall)
        .filter(ACOnCall.config_id == config.id)
        .options(selectinload(ACOnCall.ac))
    )
    acs_oc = result.scalars().all()

    result = await db.execute(
        select(DutyAssignment)
        .filter(
            DutyAssignment.config_id == config.id,
            DutyAssignment.duty_type.in_(
                [DutyType.WARD_MO, DutyType.EOT_MO, DutyType.EOT]
            ),
        )
        .options(selectinload(DutyAssignment.staff))
    )
    duties = result.scalars().all()

    consultant_by_date = {c.date: c for c in consultants_oc}
    ac_by_date = {a.date: a for a in acs_oc}
    call_by_date_type: dict[tuple, list] = defaultdict(list)
    for ca in calls:
        call_by_date_type[(ca.date, ca.call_type)].append(ca)

    ward_mo_by_date: dict[date, list[str]] = defaultdict(list)
    eot_mo_by_date: dict[date, list[str]] = defaultdict(list)
    eot_by_date: dict[date, list[str]] = defaultdict(list)
    for d in duties:
        name = d.staff.name if d.staff else ""
        if d.duty_type == DutyType.WARD_MO:
            ward_mo_by_date[d.date].append(name)
        elif d.duty_type == DutyType.EOT_MO:
            eot_mo_by_date[d.date].append(name)
        elif d.duty_type == DutyType.EOT:
            eot_by_date[d.date].append(name)

    headers = (
        ["Date", "Day", "Consultant", "AC"]
        + call_type_cols
        + ["Ward MO", "EOT MO", "EOT"]
    )
    num_cols = len(headers)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    title_cell = ws.cell(
        row=1, column=1, value=f"Call Roster - {month_name}'{year_2digit}"
    )
    title_cell.font = TITLE_FONT
    title_cell.alignment = CENTER

    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=3, column=col_idx, value=header)
    _style_header_row(ws, 3, num_cols)

    day_abbrevs = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]

    for day in range(1, num_days + 1):
        row = day + 3
        current_date = date(year, month, day)
        weekday = current_date.weekday()
        is_weekend = weekday >= 5
        is_ph = current_date in ph_dates

        day_label = day_abbrevs[weekday]
        if is_ph:
            day_label += " (PH)"

        cons_label = ""
        if current_date in consultant_by_date:
            coc = consultant_by_date[current_date]
            cons_name = coc.consultant.name if coc.consultant else ""
            if coc.supervising_consultant_id and coc.supervising_consultant:
                cons_label = f"{cons_name} / {coc.supervising_consultant.name}"
            else:
                cons_label = cons_name

        ac_entry = ac_by_date.get(current_date)
        ac_label = ac_entry.ac.name if ac_entry and ac_entry.ac else ""

        row_values = [day, day_label, cons_label, ac_label]
        call_overrides: list[bool] = []
        for ct_name in call_type_cols:
            assignments = call_by_date_type.get((current_date, ct_name), [])
            names = ", ".join(a.staff.name for a in assignments if a.staff)
            is_override = any(a.is_manual_override for a in assignments)
            row_values.append(names)
            call_overrides.append(is_override)

        row_values.append(", ".join(ward_mo_by_date.get(current_date, [])))
        row_values.append(", ".join(eot_mo_by_date.get(current_date, [])))
        row_values.append(", ".join(eot_by_date.get(current_date, [])))

        for col_idx, value in enumerate(row_values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = CENTER
            if is_ph:
                cell.fill = PH_FILL
            elif is_weekend:
                cell.fill = WEEKEND_FILL

        # Call type cols start at column 5 (after Date, Day, Consultant, AC)
        for i, is_override in enumerate(call_overrides):
            if is_override:
                ws.cell(row=row, column=5 + i).fill = OVERRIDE_FILL

    col = 1
    col = _set_col_width(ws, col, 5)  # Date
    col = _set_col_width(ws, col, 7)  # Day
    col = _set_col_width(ws, col, 16)  # Consultant
    col = _set_col_width(ws, col, 14)  # AC
    for _ in call_type_cols:
        col = _set_col_width(ws, col, 13)
    col = _set_col_width(ws, col, 14)  # Ward MO
    col = _set_col_width(ws, col, 14)  # EOT MO
    col = _set_col_width(ws, col, 14)  # EOT


async def _build_clinic_sheet_clean(
    wb: Workbook,
    config: MonthlyConfig,
    db: AsyncSession,
    year: int,
    month: int,
    num_days: int,
    month_name: str,
):
    ws = wb.create_sheet("Clinic")
    year_2digit = str(year)[-2:]
    _apply_print_settings(ws)

    ph_dates = await _get_ph_dates(db, year, month)

    result = await db.execute(
        select(DutyAssignment)
        .filter(
            DutyAssignment.config_id == config.id,
            DutyAssignment.duty_type == DutyType.CLINIC,
        )
        .options(selectinload(DutyAssignment.staff))
        .order_by(DutyAssignment.date, DutyAssignment.session, DutyAssignment.location)
    )
    clinic_duties = result.scalars().all()

    clinic_cols: list[str] = []
    clinic_cols_seen: set[str] = set()
    for d in clinic_duties:
        key = d.clinic_type if d.clinic_type else (d.location if d.location else "?")
        if key not in clinic_cols_seen:
            clinic_cols.append(key)
            clinic_cols_seen.add(key)

    clinic_by_ds: dict[tuple, dict[str, list[str]]] = defaultdict(
        lambda: defaultdict(list)
    )
    ds_ordered: list[tuple] = []
    ds_seen: set[tuple] = set()
    for d in clinic_duties:
        key = d.clinic_type if d.clinic_type else (d.location if d.location else "?")
        name = d.staff.name if d.staff else ""
        ds = (d.date, d.session)
        if ds not in ds_seen:
            ds_ordered.append(ds)
            ds_seen.add(ds)
        clinic_by_ds[ds][key].append(name)

    headers = ["Date", "Day", "Session"] + clinic_cols
    num_cols = len(headers)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(num_cols, 1))
    title_cell = ws.cell(row=1, column=1, value=f"Clinic - {month_name}'{year_2digit}")
    title_cell.font = TITLE_FONT
    title_cell.alignment = CENTER

    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=3, column=col_idx, value=header)
    _style_header_row(ws, 3, num_cols)

    day_abbrevs = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]

    for data_row_idx, (current_date, session) in enumerate(ds_ordered):
        row = data_row_idx + 4
        weekday = current_date.weekday()
        is_ph = current_date in ph_dates
        day_label = day_abbrevs[weekday]
        if is_ph:
            day_label += " (PH)"
        session_label = session.value if hasattr(session, "value") else str(session)

        row_values = [current_date.day, day_label, session_label]
        for ck in clinic_cols:
            names = clinic_by_ds[(current_date, session)].get(ck, [])
            row_values.append(", ".join(names) if names else "")

        for col_idx, value in enumerate(row_values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = CENTER
            cell.fill = PH_FILL if is_ph else CLINIC_FILL

    col = 1
    col = _set_col_width(ws, col, 5)  # Date
    col = _set_col_width(ws, col, 7)  # Day
    col = _set_col_width(ws, col, 8)  # Session
    for _ in clinic_cols:
        col = _set_col_width(ws, col, 12)


async def _build_combined_call_ot(
    wb: Workbook,
    config: MonthlyConfig,
    db: AsyncSession,
    year: int,
    month: int,
    num_days: int,
    month_name: str,
):
    ws = wb.create_sheet("Call + OT")
    year_2digit = str(year)[-2:]
    _apply_print_settings(ws)

    call_type_cols = await _get_call_type_columns(db)
    ph_dates = await _get_ph_dates(db, year, month)

    result = await db.execute(
        select(CallAssignment)
        .filter(CallAssignment.config_id == config.id)
        .options(selectinload(CallAssignment.staff))
    )
    calls = result.scalars().all()

    result = await db.execute(
        select(ConsultantOnCall)
        .filter(ConsultantOnCall.config_id == config.id)
        .options(
            selectinload(ConsultantOnCall.consultant),
            selectinload(ConsultantOnCall.supervising_consultant),
        )
    )
    consultants_oc = result.scalars().all()

    result = await db.execute(
        select(ACOnCall)
        .filter(ACOnCall.config_id == config.id)
        .options(selectinload(ACOnCall.ac))
    )
    acs_oc = result.scalars().all()

    result = await db.execute(
        select(DutyAssignment)
        .filter(
            DutyAssignment.config_id == config.id,
            DutyAssignment.duty_type.in_(
                [
                    DutyType.OT,
                    DutyType.WARD_MO,
                    DutyType.EOT_MO,
                    DutyType.EOT,
                    DutyType.SPECIAL,
                ]
            ),
        )
        .options(selectinload(DutyAssignment.staff))
    )
    duties = result.scalars().all()

    consultant_by_date = {c.date: c for c in consultants_oc}
    ac_by_date = {a.date: a for a in acs_oc}
    call_by_date_type: dict[tuple, list] = defaultdict(list)
    for ca in calls:
        call_by_date_type[(ca.date, ca.call_type)].append(ca)

    ot_by_date: dict[date, dict[str, list[str]]] = defaultdict(
        lambda: defaultdict(list)
    )
    ot_locations: list[str] = []
    ot_locations_seen: set[str] = set()
    ward_mo_by_date: dict[date, list[str]] = defaultdict(list)
    eot_mo_by_date: dict[date, list[str]] = defaultdict(list)
    eot_by_date: dict[date, list[str]] = defaultdict(list)
    special_by_date: dict[date, list[str]] = defaultdict(list)

    for d in duties:
        name = d.staff.name if d.staff else ""
        if d.duty_type == DutyType.OT:
            loc = d.location or "?"
            if loc not in ot_locations_seen:
                ot_locations.append(loc)
                ot_locations_seen.add(loc)
            ot_by_date[d.date][loc].append(name)
        elif d.duty_type == DutyType.WARD_MO:
            ward_mo_by_date[d.date].append(name)
        elif d.duty_type == DutyType.EOT_MO:
            eot_mo_by_date[d.date].append(name)
        elif d.duty_type == DutyType.EOT:
            eot_by_date[d.date].append(name)
        elif d.duty_type == DutyType.SPECIAL:
            special_by_date[d.date].append(name)

    ot_locations.sort()

    headers = (
        ["Date", "Day", "Consultant", "AC"]
        + call_type_cols
        + ["Ward MO", "EOT MO", "Special", "EOT"]
        + ot_locations
    )
    num_cols = len(headers)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    title_cell = ws.cell(
        row=1, column=1, value=f"Call + OT Roster - {month_name}'{year_2digit}"
    )
    title_cell.font = TITLE_FONT
    title_cell.alignment = CENTER

    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=3, column=col_idx, value=header)
    _style_header_row(ws, 3, num_cols)

    day_abbrevs = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]

    for day in range(1, num_days + 1):
        row = day + 3
        current_date = date(year, month, day)
        weekday = current_date.weekday()
        is_weekend = weekday >= 5
        is_ph = current_date in ph_dates

        day_label = day_abbrevs[weekday]
        if is_ph:
            day_label += " (PH)"

        cons_label = ""
        if current_date in consultant_by_date:
            coc = consultant_by_date[current_date]
            cons_name = coc.consultant.name if coc.consultant else ""
            if coc.supervising_consultant_id and coc.supervising_consultant:
                cons_label = f"{cons_name} / {coc.supervising_consultant.name}"
            else:
                cons_label = cons_name

        ac_entry = ac_by_date.get(current_date)
        ac_label = ac_entry.ac.name if ac_entry and ac_entry.ac else ""

        row_values = [day, day_label, cons_label, ac_label]
        call_overrides: list[bool] = []
        for ct_name in call_type_cols:
            assignments = call_by_date_type.get((current_date, ct_name), [])
            names = ", ".join(a.staff.name for a in assignments if a.staff)
            is_override = any(a.is_manual_override for a in assignments)
            row_values.append(names)
            call_overrides.append(is_override)

        row_values.append(", ".join(ward_mo_by_date.get(current_date, [])))
        row_values.append(", ".join(eot_mo_by_date.get(current_date, [])))
        row_values.append(", ".join(special_by_date.get(current_date, [])))
        row_values.append(", ".join(eot_by_date.get(current_date, [])))
        for loc in ot_locations:
            row_values.append(", ".join(ot_by_date[current_date].get(loc, [])))

        for col_idx, value in enumerate(row_values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = CENTER
            if is_ph:
                cell.fill = PH_FILL
            elif is_weekend:
                cell.fill = WEEKEND_FILL

        # Call type cols start at column 5 (after Date, Day, Consultant, AC)
        for i, is_override in enumerate(call_overrides):
            if is_override:
                ws.cell(row=row, column=5 + i).fill = OVERRIDE_FILL

    col = 1
    col = _set_col_width(ws, col, 5)  # Date
    col = _set_col_width(ws, col, 7)  # Day
    col = _set_col_width(ws, col, 16)  # Consultant
    col = _set_col_width(ws, col, 14)  # AC
    for _ in call_type_cols:
        col = _set_col_width(ws, col, 13)
    col = _set_col_width(ws, col, 12)  # Ward MO
    col = _set_col_width(ws, col, 12)  # EOT MO
    col = _set_col_width(ws, col, 12)  # Special
    col = _set_col_width(ws, col, 12)  # EOT
    for _ in ot_locations:
        col = _set_col_width(ws, col, 14)


async def _build_combined_ot_clinic(
    wb: Workbook,
    config: MonthlyConfig,
    db: AsyncSession,
    year: int,
    month: int,
    num_days: int,
    month_name: str,
):
    ws = wb.create_sheet("OT + Clinic")
    year_2digit = str(year)[-2:]
    _apply_print_settings(ws)

    ph_dates = await _get_ph_dates(db, year, month)

    result = await db.execute(
        select(DutyAssignment)
        .filter(
            DutyAssignment.config_id == config.id,
            DutyAssignment.duty_type == DutyType.OT,
        )
        .options(selectinload(DutyAssignment.staff))
        .order_by(DutyAssignment.date, DutyAssignment.session, DutyAssignment.location)
    )
    ot_duties = result.scalars().all()

    result = await db.execute(
        select(DutyAssignment)
        .filter(
            DutyAssignment.config_id == config.id,
            DutyAssignment.duty_type == DutyType.CLINIC,
        )
        .options(selectinload(DutyAssignment.staff))
        .order_by(DutyAssignment.date, DutyAssignment.session, DutyAssignment.location)
    )
    clinic_duties = result.scalars().all()

    ot_locations: list[str] = []
    ot_locations_seen: set[str] = set()
    for d in ot_duties:
        loc = d.location or "?"
        if loc not in ot_locations_seen:
            ot_locations.append(loc)
            ot_locations_seen.add(loc)

    ot_locations.sort()

    clinic_cols: list[str] = []
    clinic_cols_seen: set[str] = set()
    for d in clinic_duties:
        key = d.clinic_type if d.clinic_type else (d.location if d.location else "?")
        if key not in clinic_cols_seen:
            clinic_cols.append(key)
            clinic_cols_seen.add(key)

    ot_by_ds: dict[tuple, dict[str, list[str]]] = defaultdict(lambda: defaultdict(list))
    for d in ot_duties:
        ot_by_ds[(d.date, d.session)][d.location or "?"].append(
            d.staff.name if d.staff else ""
        )

    clinic_by_ds: dict[tuple, dict[str, list[str]]] = defaultdict(
        lambda: defaultdict(list)
    )
    for d in clinic_duties:
        key = d.clinic_type if d.clinic_type else (d.location if d.location else "?")
        clinic_by_ds[(d.date, d.session)][key].append(d.staff.name if d.staff else "")

    # Merge all (date, session) pairs and sort by date then session
    session_order = {"AM": 0, "PM": 1, "Full Day": 2}
    all_ds: set[tuple] = set(ot_by_ds.keys()) | set(clinic_by_ds.keys())
    ds_ordered = sorted(
        all_ds,
        key=lambda ds: (
            ds[0],
            session_order.get(
                ds[1].value if hasattr(ds[1], "value") else str(ds[1]), 99
            ),
        ),
    )

    headers = ["Date", "Day", "Session"] + ot_locations + clinic_cols
    num_cols = len(headers)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(num_cols, 1))
    title_cell = ws.cell(
        row=1, column=1, value=f"OT + Clinic Roster - {month_name}'{year_2digit}"
    )
    title_cell.font = TITLE_FONT
    title_cell.alignment = CENTER

    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=3, column=col_idx, value=header)
    _style_header_row(ws, 3, num_cols)

    day_abbrevs = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]

    for data_row_idx, (current_date, session) in enumerate(ds_ordered):
        row = data_row_idx + 4
        weekday = current_date.weekday()
        is_weekend = weekday >= 5
        is_ph = current_date in ph_dates

        day_label = day_abbrevs[weekday]
        if is_ph:
            day_label += " (PH)"
        session_label = session.value if hasattr(session, "value") else str(session)

        row_values = [current_date.day, day_label, session_label]
        for loc in ot_locations:
            names = ot_by_ds[(current_date, session)].get(loc, [])
            row_values.append(", ".join(names) if names else "")
        for ck in clinic_cols:
            names = clinic_by_ds[(current_date, session)].get(ck, [])
            row_values.append(", ".join(names) if names else "")

        # Fixed prefix cols (Date, Day, Session) + OT room cols use weekend/PH fill.
        # Clinic cols get CLINIC_FILL; PH overrides everything.
        ot_section_end = 3 + len(ot_locations)  # last col index of OT section (1-based)
        for col_idx, value in enumerate(row_values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = CENTER
            if is_ph:
                cell.fill = PH_FILL
            elif col_idx > ot_section_end:
                cell.fill = CLINIC_FILL
            elif is_weekend:
                cell.fill = WEEKEND_FILL

    col = 1
    col = _set_col_width(ws, col, 5)  # Date
    col = _set_col_width(ws, col, 7)  # Day
    col = _set_col_width(ws, col, 8)  # Session
    for _ in ot_locations:
        col = _set_col_width(ws, col, 14)
    for _ in clinic_cols:
        col = _set_col_width(ws, col, 12)
